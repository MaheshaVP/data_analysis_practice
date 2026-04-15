#excel file handling

# from openpyxl import Workbook

# wb = Workbook()
# ws = wb.active
# ws.title = "My_sheet"

# ws.append(['name','age','city'])
# ws.append(['Mahesh',22,'Pavagada'])
# ws.append(['Sheela',22,'Bangalore'])

# wb.save("my_sheet.xlsx")

from openpyxl import load_workbook

wb = load_workbook('my_sheet.xlsx')
sheet = wb.active

for row in sheet.iter_rows(values_only=True):
    print(row)
